from openpyxl import load_workbook, Workbook

# originalWb = load_workbook(filename="dataset.xlsx")

# ws = originalWb.active

newWb = Workbook()

newWb.create_sheet(title="test")
newWb.active = newWb["test"]


class Main:

    def __init__(self):
        self.oldWb = load_workbook(filename="dataset.xlsx")
        self.newWb = Workbook()
        self.newWb.create_sheet("country")


    def run(self):

        oldSource = self.oldWb["data"]

        for cell in oldSource["Nationality"]:
            print(cell)


        newWb.save(filename="dataset_changed.xlsx")



if __name__ == "__main__":
    main = Main()
    main.run()